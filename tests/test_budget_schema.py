"""Day 3. Tests for parse_inputs across long, wide and edge cases."""
from __future__ import annotations

from io import BytesIO

import pytest
from budget_schema import (
    DATA_COLS,
    MAX_ROWS,
    parse_inputs,
)
from openpyxl import Workbook


def _new_wb() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _write_metadata(wb: Workbook, *, currency: str = "GBP") -> None:
    ws = wb.create_sheet("metadata")
    ws.append(["key", "value"])
    ws.append(["company", "TestCo Ltd"])
    ws.append(["currency", currency])
    ws.append(["fiscal_year", "FY26"])
    ws.append(["period_label", "April 2026"])


def _write_data_long(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("data")
    headers = list(DATA_COLS)
    ws.append(headers)
    for r in rows:
        ws.append([r.get(c) for c in headers])


def _to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _basic_row(**overrides) -> dict:
    base = {
        "period": "2026-04",
        "cost_centre": "Engineering",
        "category": "Salaries",
        "line_type": "cost",
        "budget": 1000,
        "actual": 1100,
        "tolerance_pct": None,
        "notes": None,
    }
    base.update(overrides)
    return base


# ---- Long format happy paths ------------------------------------------

class TestLongHappyPath:
    def test_minimal_workbook(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row()])
        result = parse_inputs(file_bytes=_to_bytes(wb))
        assert result.metadata.company == "TestCo Ltd"
        assert result.metadata.currency == "GBP"
        assert len(result.rows) == 1
        assert result.rows[0].budget == 1000
        assert result.rows[0].actual == 1100
        assert result.warnings == []

    def test_line_type_defaults_to_cost(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row(line_type=None)])
        result = parse_inputs(file_bytes=_to_bytes(wb))
        assert result.rows[0].line_type == "cost"

    def test_revenue_lines_accepted(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row(line_type="revenue")])
        assert parse_inputs(file_bytes=_to_bytes(wb)).rows[0].line_type == "revenue"

    def test_optional_tolerance_and_notes(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row(tolerance_pct=0.15, notes="seasonality")])
        r = parse_inputs(file_bytes=_to_bytes(wb)).rows[0]
        assert r.tolerance_pct == 0.15
        assert r.notes == "seasonality"

    def test_missing_actual_treated_as_zero_with_warning(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row(actual=None)])
        result = parse_inputs(file_bytes=_to_bytes(wb))
        assert result.rows[0].actual == 0.0
        assert any("In-progress" in w for w in result.warnings)

    def test_all_zero_rows_dropped_with_warning(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [
            _basic_row(),
            _basic_row(category="Software", budget=0, actual=0),
            _basic_row(category="Travel",   budget=0, actual=0),
        ])
        result = parse_inputs(file_bytes=_to_bytes(wb))
        assert len(result.rows) == 1
        assert any("Dropped 2" in w for w in result.warnings)


# ---- Wide format adapter ----------------------------------------------

class TestWideAdapter:
    def _wide_wb(self) -> Workbook:
        wb = _new_wb()
        _write_metadata(wb)
        b = wb.create_sheet("budget")
        a = wb.create_sheet("actual")
        b.append(["period", "cost_centre", "category", "line_type", "amount"])
        a.append(["period", "cost_centre", "category", "line_type", "amount"])
        b.append(["2026-04", "Eng", "Salaries", "cost", 1000])
        a.append(["2026-04", "Eng", "Salaries", "cost", 1100])
        b.append(["2026-04", "Eng", "Software", "cost", 200])
        a.append(["2026-04", "Eng", "Software", "cost", 250])
        return wb

    def test_inner_join_produces_long(self):
        result = parse_inputs(file_bytes=_to_bytes(self._wide_wb()))
        assert len(result.rows) == 2
        sal = next(r for r in result.rows if r.category == "Salaries")
        assert sal.budget == 1000 and sal.actual == 1100

    def test_actual_only_row_dropped(self):
        wb = self._wide_wb()
        wb["actual"].append(["2026-04", "Mktg", "Travel", "cost", 50])
        result = parse_inputs(file_bytes=_to_bytes(wb))
        # Mktg is not in budget so not in the inner join.
        assert all(r.cost_centre != "Mktg" for r in result.rows)

    def test_budget_only_row_keeps_actual_zero(self):
        wb = self._wide_wb()
        wb["budget"].append(["2026-04", "Mktg", "Travel", "cost", 50])
        result = parse_inputs(file_bytes=_to_bytes(wb))
        mktg = next(r for r in result.rows if r.cost_centre == "Mktg")
        assert mktg.budget == 50
        assert mktg.actual == 0
        # And it shows up in warnings as in-progress.
        assert any("In-progress" in w for w in result.warnings)


# ---- Rejection cases --------------------------------------------------

class TestRejection:
    def test_both_data_and_wide_rejected(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row()])
        wb.create_sheet("budget").append(["period", "cost_centre", "category", "amount"])
        wb.create_sheet("actual").append(["period", "cost_centre", "category", "amount"])
        with pytest.raises(ValueError, match="both layouts"):
            parse_inputs(file_bytes=_to_bytes(wb))

    def test_missing_required_column(self):
        wb = _new_wb()
        _write_metadata(wb)
        ws = wb.create_sheet("data")
        # missing 'actual'
        ws.append(["period", "cost_centre", "category", "line_type", "budget"])
        ws.append(["2026-04", "Eng", "Salaries", "cost", 1000])
        with pytest.raises(ValueError, match="missing column"):
            parse_inputs(file_bytes=_to_bytes(wb))

    def test_missing_metadata_sheet(self):
        wb = _new_wb()
        _write_data_long(wb, [_basic_row()])
        with pytest.raises(ValueError, match="metadata"):
            parse_inputs(file_bytes=_to_bytes(wb))

    def test_non_gbp_currency_rejected(self):
        wb = _new_wb()
        _write_metadata(wb, currency="EUR")
        _write_data_long(wb, [_basic_row()])
        with pytest.raises(ValueError, match="GBP only"):
            parse_inputs(file_bytes=_to_bytes(wb))

    def test_duplicate_key_rejected(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [
            _basic_row(),
            _basic_row(),  # same period+cost_centre+category+line_type
        ])
        with pytest.raises(ValueError, match="Duplicate row"):
            parse_inputs(file_bytes=_to_bytes(wb))

    def test_missing_budget_value_rejected(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row(budget=None)])
        with pytest.raises(ValueError, match="budget"):
            parse_inputs(file_bytes=_to_bytes(wb))

    def test_invalid_line_type_rejected(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row(line_type="capex")])
        with pytest.raises(ValueError, match="line_type"):
            parse_inputs(file_bytes=_to_bytes(wb))

    def test_metadata_missing_keys_rejected(self):
        wb = _new_wb()
        ws = wb.create_sheet("metadata")
        ws.append(["key", "value"])
        ws.append(["company", "X"])
        # Missing currency, fiscal_year, period_label
        _write_data_long(wb, [_basic_row()])
        with pytest.raises(ValueError, match="metadata"):
            parse_inputs(file_bytes=_to_bytes(wb))


# ---- Cap and sign warnings -------------------------------------------

class TestCapAndSigns:
    def test_row_cap_warns(self):
        wb = _new_wb()
        _write_metadata(wb)
        rows = []
        # Make MAX_ROWS + 5 unique rows.
        for i in range(MAX_ROWS + 5):
            rows.append(_basic_row(cost_centre=f"CC{i:04d}"))
        _write_data_long(wb, rows)
        result = parse_inputs(file_bytes=_to_bytes(wb))
        assert len(result.rows) == MAX_ROWS
        assert any("truncated" in w for w in result.warnings)

    def test_negative_revenue_warns_but_does_not_flip(self):
        wb = _new_wb()
        _write_metadata(wb)
        _write_data_long(wb, [_basic_row(line_type="revenue", budget=-100, actual=-90)])
        result = parse_inputs(file_bytes=_to_bytes(wb))
        assert result.rows[0].budget == -100
        assert any("negative value" in w for w in result.warnings)
