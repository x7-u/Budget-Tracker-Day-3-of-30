"""Day 3. Tests for the Excel writer."""
from __future__ import annotations

from pathlib import Path

from budget_schema import BudgetMetadata, BudgetRow
from excel_writer import RAG_FILL, write_workbook
from openpyxl import load_workbook
from pipeline import AnalysisResult, Commentary
from variance import (
    aggregate_by,
    compute_all,
    headline_stats,
)


def _row(**overrides) -> BudgetRow:
    base = dict(
        period="2026-04",
        cost_centre="Engineering",
        category="Salaries",
        line_type="cost",
        budget=1000.0,
        actual=1100.0,
        tolerance_pct=None,
        notes=None,
    )
    base.update(overrides)
    return BudgetRow(**base)


def _result(rows: list[BudgetRow], *, periods: list[str] | None = None) -> AnalysisResult:
    if periods:
        rows = [_row(**{**r.__dict__, "period": p}) for p, r in zip(periods, rows, strict=False)]
    var_rows = compute_all(rows)
    return AnalysisResult(
        metadata=BudgetMetadata(company="TestCo", currency="GBP",
                                fiscal_year="FY26", period_label="April 2026"),
        rows=var_rows,
        by_cost_centre=aggregate_by(var_rows, dim="cost_centre"),
        by_category=aggregate_by(var_rows, dim="category"),
        by_period=aggregate_by(var_rows, dim="period"),
        headline=headline_stats(var_rows),
        commentary=Commentary(skipped=True),
        warnings=[],
        source_filename="unit-test.xlsx",
        elapsed_ms=0,
    )


def test_single_period_workbook_has_six_sheets(tmp_path: Path) -> None:
    rows = [_row(category="Salaries"), _row(category="Software", actual=1300)]
    out = write_workbook(_result(rows), tmp_path)
    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Summary", "By Cost Centre", "By Category", "Detail", "Commentary", "Inputs",
    ]


def test_multi_period_workbook_has_seven_sheets(tmp_path: Path) -> None:
    rows = [
        _row(period="2026-Q1"),
        _row(period="2026-Q2", category="Software"),
        _row(period="2026-Q3", category="Travel"),
    ]
    out = write_workbook(_result(rows), tmp_path)
    wb = load_workbook(out)
    assert "By Period" in wb.sheetnames
    assert len(wb.sheetnames) == 7


def test_detail_has_autofilter(tmp_path: Path) -> None:
    rows = [_row(category="Salaries"), _row(category="Software", actual=1300)]
    out = write_workbook(_result(rows), tmp_path)
    wb = load_workbook(out)
    detail = wb["Detail"]
    assert detail.auto_filter.ref is not None
    assert detail.auto_filter.ref.startswith("A1")


def test_detail_row_count_matches_input(tmp_path: Path) -> None:
    rows = [
        _row(category="Salaries"),
        _row(category="Software", actual=1300),
        _row(category="Travel",  actual=900),
    ]
    out = write_workbook(_result(rows), tmp_path)
    wb = load_workbook(out)
    detail = wb["Detail"]
    # header + 3 rows = 4
    assert detail.max_row == 4


def test_rag_cell_uses_palette_fill(tmp_path: Path) -> None:
    rows = [_row(category="Salaries", actual=1300)]   # +30%, red
    out = write_workbook(_result(rows), tmp_path)
    wb = load_workbook(out)
    detail = wb["Detail"]
    rag_cell = detail.cell(row=2, column=11)
    assert rag_cell.value == "RED"
    expected = RAG_FILL["red"].fgColor.value
    actual = rag_cell.fill.fgColor.value
    # openpyxl returns ARGB on read; strip the leading FF if present.
    assert (actual or "").endswith(expected)


def test_currency_format_string_applied(tmp_path: Path) -> None:
    out = write_workbook(_result([_row()]), tmp_path)
    wb = load_workbook(out)
    detail = wb["Detail"]
    assert "£" in detail.cell(row=2, column=5).number_format


def test_filename_includes_company_period_and_timestamp(tmp_path: Path) -> None:
    out = write_workbook(_result([_row()]), tmp_path)
    name = out.name
    assert name.startswith("budget_variance_")
    assert "testco" in name
    assert "april-2026" in name
    assert name.endswith(".xlsx")


def test_handles_zero_budget_row_with_na_label(tmp_path: Path) -> None:
    out = write_workbook(_result([_row(budget=0.0, actual=100.0)]), tmp_path)
    wb = load_workbook(out)
    detail = wb["Detail"]
    pct_cell = detail.cell(row=2, column=8)
    assert pct_cell.value == "n/a"
    rag_cell = detail.cell(row=2, column=11)
    assert rag_cell.value == "NA"
