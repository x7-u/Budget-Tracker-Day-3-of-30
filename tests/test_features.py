"""Day 3. Tests for the second-wave features.

Covers:
* Forecast column carried through into VarianceRow and forecast_variance.
* Parent rollups via aggregate_by_parent.
* Cost log JSONL append + summary.
* History store + z-score scoring.
* Comparison module diffs.
* PowerPoint writer happy path (only when python-pptx is installed).
"""
from __future__ import annotations

from pathlib import Path

import pytest
from budget_schema import BudgetMetadata, BudgetRow
from comparison import compare
from comparison import write_csv as compare_write_csv
from comparison import write_workbook as compare_write_workbook
from cost_log import CostLog
from history_store import HistoryStore, score_rows
from pipeline import AnalysisResult, Commentary, analyse
from variance import (
    aggregate_by,
    aggregate_by_parent,
    classify_row,
    compute_all,
    headline_stats,
)


def _row(**overrides) -> BudgetRow:
    base = dict(
        period="2026-04",
        cost_centre="Engineering",
        category="Salaries",
        line_type="cost",
        budget=1000.0,
        actual=1000.0,
        tolerance_pct=None,
        notes=None,
        forecast=None,
        parent=None,
    )
    base.update(overrides)
    return BudgetRow(**base)


def _result(rows, source="t.xlsx") -> AnalysisResult:
    var_rows = compute_all(rows)
    return AnalysisResult(
        metadata=BudgetMetadata(company="TestCo", currency="GBP",
                                fiscal_year="FY26", period_label="April 2026"),
        rows=var_rows,
        by_cost_centre=aggregate_by(var_rows, dim="cost_centre"),
        by_category=aggregate_by(var_rows, dim="category"),
        by_period=aggregate_by(var_rows, dim="period"),
        by_parent=aggregate_by_parent(var_rows),
        headline=headline_stats(var_rows),
        commentary=Commentary(skipped=True),
        warnings=[],
        source_filename=source,
        elapsed_ms=0,
    )


# ---- Forecast column --------------------------------------------------

class TestForecast:
    def test_forecast_carries_through(self):
        v = classify_row(_row(forecast=1050.0, actual=1080.0))
        assert v.forecast == 1050.0
        assert v.forecast_variance == 30.0
        assert v.forecast_variance_pct == pytest.approx(30 / 1050)

    def test_forecast_none_yields_none_variance(self):
        v = classify_row(_row(actual=1100.0))
        assert v.forecast is None
        assert v.forecast_variance is None
        assert v.forecast_variance_pct is None

    def test_zero_forecast_keeps_pct_none(self):
        v = classify_row(_row(forecast=0.0, actual=10.0))
        assert v.forecast == 0.0
        assert v.forecast_variance == 10.0
        assert v.forecast_variance_pct is None


# ---- Parent rollups ---------------------------------------------------

class TestParent:
    def test_aggregate_by_parent_groups_correctly(self):
        rows = compute_all([
            _row(cost_centre="Eng",  parent="Engineering", actual=1100),
            _row(cost_centre="Plat", parent="Engineering", actual=900,  budget=1000),
            _row(cost_centre="Sales", parent="Commercial", actual=1500, budget=1000),
        ])
        aggs = aggregate_by_parent(rows)
        keys = {a.key for a in aggs}
        assert keys == {"Engineering", "Commercial"}
        eng = next(a for a in aggs if a.key == "Engineering")
        assert eng.budget == 2000
        assert eng.actual == 2000

    def test_no_parent_returns_empty(self):
        rows = compute_all([_row()])
        assert aggregate_by_parent(rows) == []

    def test_partial_parent_only_groups_those_with_parent(self):
        rows = compute_all([
            _row(cost_centre="A", parent="P1", budget=100, actual=100),
            _row(cost_centre="B", parent=None, budget=200, actual=200),
        ])
        aggs = aggregate_by_parent(rows)
        assert len(aggs) == 1
        assert aggs[0].key == "P1"


# ---- Cost log ---------------------------------------------------------

class TestCostLog:
    def test_append_and_summary(self, tmp_path: Path):
        log = CostLog(tmp_path / "runs.jsonl")
        log.append(company="A", period_label="P1", rows=10, cost_usd=0.005,
                   model="m1", skipped=False, elapsed_ms=120, source_filename="a.xlsx")
        log.append(company="A", period_label="P2", rows=20, cost_usd=0.008,
                   model="m1", skipped=False, elapsed_ms=130)
        s = log.summary()
        assert s.runs == 2
        assert s.cost_usd_total == pytest.approx(0.013)
        assert s.rows_total == 30
        assert s.last_run_at is not None

    def test_clear_resets(self, tmp_path: Path):
        log = CostLog(tmp_path / "runs.jsonl")
        log.append(company="A", period_label="P", rows=1, cost_usd=0.001,
                   model="m", skipped=False, elapsed_ms=1)
        n = log.clear()
        assert n == 1
        assert log.summary().runs == 0


# ---- History store + z-score -----------------------------------------

class TestHistory:
    def test_score_marks_outliers_after_three_obs(self, tmp_path: Path):
        store = HistoryStore(tmp_path / "h.jsonl")
        # Seed with three historical observations all at 5% overspend.
        for _ in range(3):
            store.append_run(
                company="X", period_label="seed",
                rows=compute_all([_row(actual=1050)]),
            )
        # Now classify a runaway 30% overspend.
        live = compute_all([_row(actual=1300)])
        buckets = store.keys_with_enough_history(company="X")
        flagged = score_rows(live, buckets)
        assert flagged == 1
        assert live[0].z_score is not None
        assert abs(live[0].z_score) >= 2

    def test_score_does_not_fire_with_too_little_history(self, tmp_path: Path):
        store = HistoryStore(tmp_path / "h.jsonl")
        store.append_run(company="X", period_label="seed",
                         rows=compute_all([_row(actual=1050)]))
        live = compute_all([_row(actual=1300)])
        buckets = store.keys_with_enough_history(company="X")
        flagged = score_rows(live, buckets)
        assert flagged == 0
        assert live[0].z_score is None

    def test_company_isolated_history(self, tmp_path: Path):
        store = HistoryStore(tmp_path / "h.jsonl")
        for _ in range(3):
            store.append_run(company="A", period_label="seed",
                             rows=compute_all([_row(actual=1050)]))
        # Different company should see no history.
        buckets = store.keys_with_enough_history(company="B")
        assert buckets == {}


# ---- Comparison ------------------------------------------------------

class TestComparison:
    def test_compare_diffs_common_added_removed(self):
        a = _result([
            _row(cost_centre="Eng",  category="Salaries", actual=1100),  # adverse
            _row(cost_centre="Eng",  category="Travel",   actual=900,  budget=1000),
        ])
        b = _result([
            _row(cost_centre="Eng",  category="Salaries", actual=1500),  # bigger adverse
            _row(cost_centre="Mktg", category="Events",   actual=500,  budget=400),  # added
        ])
        c = compare(a, b)
        assert c.n_added == 1
        assert c.n_removed == 1
        assert c.n_common == 1
        # The Salaries row should be the biggest mover and lead the list.
        first = c.rows[0]
        assert first.cost_centre == "Eng"
        assert first.category == "Salaries"
        assert first.delta_variance is not None
        assert first.delta_variance > 0

    def test_compare_zero_movers_when_workbooks_match(self):
        a = _result([_row(actual=1100)])
        b = _result([_row(actual=1100)])
        c = compare(a, b, mover_threshold=1.0)
        assert c.n_movers == 0
        assert c.delta_total_variance == 0


# ---- Pipeline + history end-to-end -----------------------------------

class TestAnalysePipelineHistory:
    def test_analyse_records_history_when_enabled(self, tmp_path: Path):
        # Build a minimal workbook in memory.
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        meta = wb.create_sheet("metadata")
        meta.append(["key", "value"])
        meta.append(["company", "RT Ltd"])
        meta.append(["currency", "GBP"])
        meta.append(["fiscal_year", "FY26"])
        meta.append(["period_label", "April 2026"])
        data = wb.create_sheet("data")
        data.append(["period", "cost_centre", "category", "line_type", "budget", "actual"])
        data.append(["2026-04", "CC", "Cat", "cost", 1000, 1100])
        buf = BytesIO()
        wb.save(buf)
        wb_bytes = buf.getvalue()

        store = HistoryStore(tmp_path / "h.jsonl")
        analyse(file_bytes=wb_bytes, source_filename="t.xlsx", skip_ai=True, history=store)
        assert store.stats()["records"] == 1


# ---- Comparison exports ----------------------------------------------

class TestComparisonExports:
    def test_csv_writer_produces_utf8_with_bom(self, tmp_path: Path):
        a = _result([_row(actual=1100)])
        b = _result([_row(actual=1300)])
        c = compare(a, b)
        out = compare_write_csv(c, tmp_path / "compare.csv")
        raw = out.read_bytes()
        assert raw.startswith(b"\xef\xbb\xbf")
        text = raw.decode("utf-8-sig")
        assert "delta_variance" in text.splitlines()[0]

    def test_xlsx_writer_has_summary_and_detail(self, tmp_path: Path):
        from openpyxl import load_workbook
        a = _result([_row(actual=1100), _row(category="Travel", actual=900, budget=1000)])
        b = _result([_row(actual=1500), _row(cost_centre="Mktg", category="Events", actual=500, budget=400)])
        c = compare(a, b)
        out = compare_write_workbook(c, tmp_path)
        wb = load_workbook(out)
        assert wb.sheetnames == ["Summary", "Detail"]
        # autofilter set
        assert wb["Detail"].auto_filter.ref is not None


# ---- Run cache -------------------------------------------------------

class TestRunCache:
    def test_save_and_get(self, tmp_path: Path):
        from run_cache import RunCache
        cache = RunCache(tmp_path / "runs")
        payload = {"hello": "world", "rows": [{"a": 1}]}
        cache.save("abc123", payload)
        assert cache.get("abc123") == payload

    def test_missing_returns_none(self, tmp_path: Path):
        from run_cache import RunCache
        cache = RunCache(tmp_path / "runs")
        assert cache.get("nope") is None

    def test_remove_and_clear(self, tmp_path: Path):
        from run_cache import RunCache
        cache = RunCache(tmp_path / "runs")
        cache.save("a", {"x": 1})
        cache.save("b", {"x": 2})
        assert cache.remove("a") is True
        assert cache.get("a") is None
        assert cache.clear() == 1
        assert cache.get("b") is None


# ---- CSV input -------------------------------------------------------

class TestCsvInput:
    def test_csv_long_format_round_trips(self, tmp_path: Path):
        from budget_schema import parse_inputs

        csv_path = tmp_path / "acme_co.csv"
        csv_path.write_text(
            "period,cost_centre,category,line_type,budget,actual,tolerance_pct,notes\n"
            "2026-04,Eng,Salaries,cost,1000,1100,,\n"
            "2026-04,Eng,Travel,cost,200,180,0.15,seasonal\n",
            encoding="utf-8-sig",
        )
        result = parse_inputs(path=csv_path, source_filename=csv_path.name)
        assert len(result.rows) == 2
        assert result.metadata.currency == "GBP"
        assert "Acme Co" in result.metadata.company
        # The CSV warning is raised so the user knows defaults were applied.
        assert any("Metadata defaults" in w for w in result.warnings)
        # Per-row tolerance preserved.
        travel = next(r for r in result.rows if r.category == "Travel")
        assert travel.tolerance_pct == 0.15
        assert travel.notes == "seasonal"

    def test_csv_missing_required_column(self, tmp_path: Path):
        from budget_schema import parse_inputs
        csv_path = tmp_path / "bad.csv"
        csv_path.write_text("period,cost_centre,category,budget\n2026-04,X,Y,100\n", encoding="utf-8")
        with pytest.raises(ValueError, match="missing column"):
            parse_inputs(path=csv_path, source_filename="bad.csv")


# ---- Power BI scaffolding -------------------------------------------

class TestPowerBI:
    def test_writes_script_and_readme(self, tmp_path: Path):
        from power_bi import write_power_bi_assets
        out = write_power_bi_assets(tmp_path)
        assert out.suffix == ".m"
        assert out.is_file()
        readme = tmp_path / "README.md"
        assert readme.is_file()
        # The script should reference the outputs folder path.
        content = out.read_text(encoding="utf-8")
        assert "OutputsFolder" in content
        assert "budget_variance_" in content


# ---- PowerPoint writer (smoke) ---------------------------------------

class TestPowerPoint:
    def test_writes_pptx_when_available(self, tmp_path: Path):
        from pptx_writer import is_available, write_pptx

        if not is_available():
            pytest.skip("python-pptx not installed")

        result = _result([
            _row(cost_centre="A", actual=1100),
            _row(cost_centre="B", category="Travel", actual=800, budget=1000),
        ])
        # Give the commentary something to render.
        result.commentary = Commentary(
            headline="Variance was +£100 across two cost centres.",
            summary="Salaries overran in A; Travel underspent in B.",
            adverse_drivers=["A salaries 10 percent over"],
            favourable_drivers=["B travel 20 percent under"],
            actions=["Confirm hire timing in A"],
            cost_usd=0.0042, model="claude-haiku-4-5",
        )
        out = write_pptx(result, tmp_path)
        assert out is not None
        assert out.exists()
        assert out.suffix == ".pptx"
        # Sanity: file is not empty.
        assert out.stat().st_size > 5000
