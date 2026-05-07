"""Day 3. Comparison mode.

Diff two AnalysisResults at the row level by the canonical four-key
tuple ``(period, cost_centre, category, line_type)``. Surfaces what
moved, what's new, and what dropped between the two runs.

Also exposes Excel and CSV writers so the diff can be exported in the
same way a single analysis can.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class CompareRow:
    period: str
    cost_centre: str
    category: str
    line_type: str
    budget_a: float | None
    budget_b: float | None
    actual_a: float | None
    actual_b: float | None
    variance_a: float | None
    variance_b: float | None
    delta_variance: float | None      # variance_b - variance_a
    rag_a: str | None
    rag_b: str | None
    status: str                       # "common" | "added" | "removed"


@dataclass
class CompareResult:
    rows: list[CompareRow]
    headline_a: dict
    headline_b: dict
    delta_total_variance: float
    n_common: int
    n_added: int
    n_removed: int
    n_movers: int                     # |delta_variance| over a small threshold


def compare(result_a, result_b, *, mover_threshold: float = 100.0) -> CompareResult:
    """Compare two AnalysisResults. ``mover_threshold`` in GBP."""
    a_idx = _index(result_a.rows)
    b_idx = _index(result_b.rows)

    keys = list(a_idx) + [k for k in b_idx if k not in a_idx]
    rows: list[CompareRow] = []
    n_common = n_added = n_removed = n_movers = 0

    for key in keys:
        a = a_idx.get(key)
        b = b_idx.get(key)
        if a and b:
            status = "common"
            n_common += 1
        elif b and not a:
            status = "added"
            n_added += 1
        else:
            status = "removed"
            n_removed += 1

        var_a = a.variance if a else None
        var_b = b.variance if b else None
        delta = (var_b - var_a) if (var_a is not None and var_b is not None) else None
        if delta is not None and abs(delta) >= mover_threshold:
            n_movers += 1

        rows.append(CompareRow(
            period=key[0], cost_centre=key[1], category=key[2], line_type=key[3],
            budget_a=a.budget if a else None,
            budget_b=b.budget if b else None,
            actual_a=a.actual if a else None,
            actual_b=b.actual if b else None,
            variance_a=var_a,
            variance_b=var_b,
            delta_variance=delta,
            rag_a=a.rag if a else None,
            rag_b=b.rag if b else None,
            status=status,
        ))

    # Order by absolute delta first (movers up top), then by status.
    rows.sort(
        key=lambda r: (
            -(abs(r.delta_variance) if r.delta_variance is not None else 0),
            r.status,
            r.cost_centre,
            r.category,
        )
    )

    return CompareResult(
        rows=rows,
        headline_a=_headline_dict(result_a),
        headline_b=_headline_dict(result_b),
        delta_total_variance=result_b.headline.total_variance - result_a.headline.total_variance,
        n_common=n_common,
        n_added=n_added,
        n_removed=n_removed,
        n_movers=n_movers,
    )


def to_dict(c: CompareResult) -> dict:
    return {
        "rows": [_row_dict(r) for r in c.rows],
        "headline_a": c.headline_a,
        "headline_b": c.headline_b,
        "delta_total_variance": c.delta_total_variance,
        "n_common": c.n_common,
        "n_added": c.n_added,
        "n_removed": c.n_removed,
        "n_movers": c.n_movers,
    }


def _index(rows) -> dict[tuple, object]:
    return {(r.period, r.cost_centre, r.category, r.line_type): r for r in rows}


def _headline_dict(result) -> dict:
    h = result.headline
    return {
        "company": result.metadata.company,
        "period_label": result.metadata.period_label,
        "total_budget": h.total_budget,
        "total_actual": h.total_actual,
        "total_variance": h.total_variance,
        "total_variance_pct": h.total_variance_pct,
        "rag_counts": dict(h.rag_counts),
        "row_count": h.row_count,
    }


def _row_dict(r: CompareRow) -> dict:
    return {
        "period": r.period,
        "cost_centre": r.cost_centre,
        "category": r.category,
        "line_type": r.line_type,
        "budget_a": r.budget_a,
        "budget_b": r.budget_b,
        "actual_a": r.actual_a,
        "actual_b": r.actual_b,
        "variance_a": r.variance_a,
        "variance_b": r.variance_b,
        "delta_variance": r.delta_variance,
        "rag_a": r.rag_a,
        "rag_b": r.rag_b,
        "status": r.status,
    }


# ---- Exports ---------------------------------------------------------

CSV_COLUMNS = (
    "status",
    "period",
    "cost_centre",
    "category",
    "line_type",
    "budget_a",
    "budget_b",
    "actual_a",
    "actual_b",
    "variance_a",
    "variance_b",
    "delta_variance",
    "rag_a",
    "rag_b",
)

GBP_FORMAT = '"£"#,##0.00;[Red]-"£"#,##0.00'
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
CENTER = Alignment(horizontal="center", vertical="center")

STATUS_FILL = {
    "common":  PatternFill("solid", fgColor="EFEAE0"),
    "added":   PatternFill("solid", fgColor="2E8540"),
    "removed": PatternFill("solid", fgColor="8E2A2A"),
}


def write_csv(diff: CompareResult, out_path: Path | str) -> Path:
    """Flat CSV of the diff. utf-8-sig so Excel auto-detects on double-click."""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for r in diff.rows:
            writer.writerow({
                "status":         r.status,
                "period":         r.period,
                "cost_centre":    r.cost_centre,
                "category":       r.category,
                "line_type":      r.line_type,
                "budget_a":       "" if r.budget_a is None else r.budget_a,
                "budget_b":       "" if r.budget_b is None else r.budget_b,
                "actual_a":       "" if r.actual_a is None else r.actual_a,
                "actual_b":       "" if r.actual_b is None else r.actual_b,
                "variance_a":     "" if r.variance_a is None else r.variance_a,
                "variance_b":     "" if r.variance_b is None else r.variance_b,
                "delta_variance": "" if r.delta_variance is None else r.delta_variance,
                "rag_a":          r.rag_a or "",
                "rag_b":          r.rag_b or "",
            })
    return out


def write_workbook(diff: CompareResult, out_dir: Path | str) -> Path:
    """Two-sheet workbook: Summary + Detail. Mirrors the analysis writer's
    autofilter pattern on the Detail sheet so the controller can slice."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    a = diff.headline_a
    b = diff.headline_b
    slug_a = _slug(a.get("company") or "a")
    slug_b = _slug(b.get("company") or "b")
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out = out_dir / f"compare_{slug_a}_vs_{slug_b}_{ts}.xlsx"

    wb = Workbook()
    _write_compare_summary(wb.active, diff)
    _write_compare_detail(wb.create_sheet("Detail"), diff)
    wb.save(out)
    return out


def _write_compare_summary(ws, diff: CompareResult) -> None:
    ws.title = "Summary"
    ws["A1"] = "Comparison summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="555555")
    ws.merge_cells("A2:E2")

    headers = ["Metric", "A", "B", "Delta"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    a = diff.headline_a
    b = diff.headline_b
    rows = [
        ("Company",             a.get("company"),        b.get("company"),        ""),
        ("Period",              a.get("period_label"),   b.get("period_label"),   ""),
        ("Total budget",        a.get("total_budget"),   b.get("total_budget"),
                                (b.get("total_budget", 0) or 0) - (a.get("total_budget", 0) or 0)),
        ("Total actual",        a.get("total_actual"),   b.get("total_actual"),
                                (b.get("total_actual", 0) or 0) - (a.get("total_actual", 0) or 0)),
        ("Total variance",      a.get("total_variance"), b.get("total_variance"),
                                diff.delta_total_variance),
        ("Row count",           a.get("row_count"),      b.get("row_count"),
                                (b.get("row_count", 0) or 0) - (a.get("row_count", 0) or 0)),
    ]
    for i, (label, va, vb, delta) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        for col, val in [(2, va), (3, vb), (4, delta)]:
            cell = ws.cell(row=i, column=col, value=val)
            if isinstance(val, (int, float)) and label != "Row count":
                cell.number_format = GBP_FORMAT

    ws["A12"] = "Diff summary"
    ws["A12"].font = Font(bold=True)
    diff_rows = [
        ("Common rows",   diff.n_common),
        ("Added in B",    diff.n_added),
        ("Removed from A", diff.n_removed),
        ("Movers",        diff.n_movers),
    ]
    for i, (label, value) in enumerate(diff_rows, start=13):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    widths = [22, 22, 22, 18, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_compare_detail(ws, diff: CompareResult) -> None:
    headers = [
        "Status", "Period", "Cost centre", "Category", "Line type",
        "Budget A", "Budget B", "Actual A", "Actual B",
        "Variance A", "Variance B", "Delta variance",
        "RAG A", "RAG B",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    for r_idx, r in enumerate(diff.rows, start=2):
        status_cell = ws.cell(row=r_idx, column=1, value=r.status.upper())
        status_cell.font = Font(bold=True, color="FFFFFF" if r.status != "common" else "555555")
        status_cell.alignment = CENTER
        if r.status in STATUS_FILL:
            status_cell.fill = STATUS_FILL[r.status]

        ws.cell(row=r_idx, column=2, value=r.period)
        ws.cell(row=r_idx, column=3, value=r.cost_centre)
        ws.cell(row=r_idx, column=4, value=r.category)
        ws.cell(row=r_idx, column=5, value=r.line_type)

        for col_idx, value in [
            (6, r.budget_a), (7, r.budget_b),
            (8, r.actual_a), (9, r.actual_b),
            (10, r.variance_a), (11, r.variance_b),
            (12, r.delta_variance),
        ]:
            cell = ws.cell(row=r_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = GBP_FORMAT

        ws.cell(row=r_idx, column=13, value=(r.rag_a or "").upper())
        ws.cell(row=r_idx, column=14, value=(r.rag_b or "").upper())

    if diff.rows:
        ws.auto_filter.ref = ws.dimensions

    widths = [10, 12, 22, 22, 10, 14, 14, 14, 14, 14, 14, 14, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _slug(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "untitled"
