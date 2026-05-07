"""Day 3. Excel writer for the budget vs actual analysis pack.

Produces an XLSX with these sheets, in order:

* Summary           Title, run timestamp, headline KPIs, RAG distribution
* By Cost Centre    Pivot with totals row, RAG-coloured rag column
* By Category       Same shape
* By Period         Only when more than one period is present
* Detail            Every variance row, RAG-coloured, Excel autofilter on
* Commentary        AI-drafted management commentary
* Inputs            Parsed input echoed back for reproducibility

The ``Detail`` sheet's autofilter is the drill-through tool. The brief
calls for slicers; in Flask we hand that responsibility to Excel itself.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from shared.chart_styles import RAG_HEX

BLUE_HEX = "1F4E79"
NA_HEX = "B6BCC6"

RAG_FILL: dict[str, PatternFill] = {
    "green":      PatternFill("solid", fgColor=RAG_HEX["green"]),
    "amber":      PatternFill("solid", fgColor=RAG_HEX["amber"]),
    "red":        PatternFill("solid", fgColor=RAG_HEX["red"]),
    "favourable": PatternFill("solid", fgColor=BLUE_HEX),
    "na":         PatternFill("solid", fgColor=NA_HEX),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

GBP_FORMAT = '"£"#,##0.00;[Red]-"£"#,##0.00'
PCT_FORMAT = "0.00%"


# ---- Public entry point ----------------------------------------------

def write_workbook(result, out_dir: Path | str) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    company = _slug(result.metadata.company)
    period = _slug(result.metadata.period_label)
    # Seconds in the timestamp so back-to-back runs in the same minute do not collide.
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out = out_dir / f"budget_variance_{company}_{period}_{ts}.xlsx"

    wb = Workbook()
    _write_summary(wb.active, result)
    if result.by_parent:
        _write_aggregate_sheet(wb.create_sheet("By Parent"),  result.by_parent,      "Parent")
    cc_sheet = wb.create_sheet("By Cost Centre")
    _write_aggregate_sheet(cc_sheet, result.by_cost_centre, "Cost Centre")
    _write_aggregate_sheet(wb.create_sheet("By Category"),    result.by_category,    "Category")
    if result.headline.period_count > 1:
        _write_aggregate_sheet(wb.create_sheet("By Period"),  result.by_period,      "Period")
    _write_detail(wb.create_sheet("Detail"), result.rows)
    _write_commentary(wb.create_sheet("Commentary"), result.commentary, result.metadata)
    _write_inputs(wb.create_sheet("Inputs"), result)

    # Chart on Summary references the By Cost Centre sheet.
    _embed_summary_chart(wb["Summary"], cc_sheet, n_rows=len(result.by_cost_centre))

    wb.save(out)
    return out


def _embed_summary_chart(summary_ws, cc_ws, *, n_rows: int) -> None:
    """Add a horizontal bar chart of variance by cost centre to the Summary sheet."""
    if n_rows == 0:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Variance by cost centre (GBP)"
    chart.y_axis.title = "Cost centre"
    chart.x_axis.title = "Variance"
    # Column 4 of the By Cost Centre sheet is Variance, column 1 is the key.
    data = Reference(cc_ws, min_col=4, min_row=1, max_row=n_rows + 1, max_col=4)
    cats = Reference(cc_ws, min_col=1, min_row=2, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = max(8, min(20, n_rows * 0.45 + 4))
    chart.width = 16
    summary_ws.add_chart(chart, "D4")


# ---- Summary sheet ---------------------------------------------------

def _write_summary(ws, result) -> None:
    ws.title = "Summary"
    ws["A1"] = f"Budget vs Actual.  {result.metadata.company}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = (
        f"Period: {result.metadata.period_label}  |  "
        f"Currency: {result.metadata.currency}  |  "
        f"Fiscal year: {result.metadata.fiscal_year}  |  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws["A2"].font = Font(italic=True, color="555555")
    ws.merge_cells("A2:F2")

    h = result.headline

    # Headline KPIs
    ws["A4"] = "Headline"
    ws["A4"].font = Font(bold=True)
    headers = ["Metric", "Value"]
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    rows: list[tuple[str, object, str | None]] = [
        ("Total budget",   h.total_budget,   GBP_FORMAT),
        ("Total actual",   h.total_actual,   GBP_FORMAT),
        ("Total variance", h.total_variance, GBP_FORMAT),
        ("Variance %",     h.total_variance_pct if h.total_variance_pct is not None else "n/a",
            PCT_FORMAT if h.total_variance_pct is not None else None),
        ("Rows",           h.row_count, None),
        ("Periods",        h.period_count, None),
        ("Adverse lines",       h.adverse_count, None),
        ("Favourable lines",    h.favourable_count, None),
        ("n/a lines (zero budget)", h.na_count, None),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        if fmt:
            cell.number_format = fmt

    # RAG distribution
    rag_start_row = 6 + len(rows) + 2
    ws.cell(row=rag_start_row - 1, column=1, value="RAG distribution").font = Font(bold=True)
    for i, name in enumerate(["Band", "Count"], start=1):
        c = ws.cell(row=rag_start_row, column=i, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    bands = ["red", "amber", "green", "favourable", "na"]
    for i, band in enumerate(bands, start=rag_start_row + 1):
        cell = ws.cell(row=i, column=1, value=band.upper())
        cell.fill = RAG_FILL[band]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = CENTER
        ws.cell(row=i, column=2, value=h.rag_counts.get(band, 0))

    widths = [28, 18, 4, 4, 4, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Aggregate sheet (By Cost Centre / Category / Period) -----------

def _write_aggregate_sheet(ws, aggregates: list, key_label: str) -> None:
    headers = [
        key_label, "Budget", "Actual", "Variance", "Variance %",
        "Red", "Amber", "Green", "Favourable", "n/a", "Rows",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    for r_idx, a in enumerate(aggregates, start=2):
        ws.cell(row=r_idx, column=1, value=a.key)
        bud = ws.cell(row=r_idx, column=2, value=a.budget)
        bud.number_format = GBP_FORMAT
        act = ws.cell(row=r_idx, column=3, value=a.actual)
        act.number_format = GBP_FORMAT
        var = ws.cell(row=r_idx, column=4, value=a.variance)
        var.number_format = GBP_FORMAT
        if a.variance_pct is not None:
            pct = ws.cell(row=r_idx, column=5, value=a.variance_pct)
            pct.number_format = PCT_FORMAT
        else:
            ws.cell(row=r_idx, column=5, value="n/a")
        ws.cell(row=r_idx, column=6,  value=a.rag_counts.get("red",        0))
        ws.cell(row=r_idx, column=7,  value=a.rag_counts.get("amber",      0))
        ws.cell(row=r_idx, column=8,  value=a.rag_counts.get("green",      0))
        ws.cell(row=r_idx, column=9,  value=a.rag_counts.get("favourable", 0))
        ws.cell(row=r_idx, column=10, value=a.rag_counts.get("na",         0))
        ws.cell(row=r_idx, column=11, value=a.row_count)

    if aggregates:
        total_row = len(aggregates) + 2
        ws.cell(row=total_row, column=1, value="Total").font = TOTAL_FONT
        for col, key in [(2, "budget"), (3, "actual"), (4, "variance")]:
            v = sum(getattr(a, key) for a in aggregates)
            cell = ws.cell(row=total_row, column=col, value=v)
            cell.number_format = GBP_FORMAT
            cell.font = TOTAL_FONT
        total_budget = sum(a.budget for a in aggregates)
        total_variance = sum(a.variance for a in aggregates)
        if total_budget:
            cell = ws.cell(row=total_row, column=5, value=total_variance / total_budget)
            cell.number_format = PCT_FORMAT
            cell.font = TOTAL_FONT
        for col in range(6, 12):
            ws.cell(row=total_row, column=col,
                    value=sum(_safe_int(getattr(a, "rag_counts", {}).get(_band_for_col(col), 0)
                                        if col < 11 else a.row_count) for a in aggregates)).font = TOTAL_FONT

    widths = [28, 14, 14, 14, 12, 8, 8, 8, 12, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _band_for_col(col: int) -> str:
    return {6: "red", 7: "amber", 8: "green", 9: "favourable", 10: "na"}[col]


def _safe_int(v) -> int:
    try:
        return int(v)
    except Exception:
        return 0


# ---- Detail sheet ---------------------------------------------------

def _write_detail(ws, rows: list) -> None:
    headers = [
        "Period", "Cost Centre", "Category", "Line type",
        "Budget", "Actual", "Variance", "Variance %",
        "Tolerance %", "Adverse", "RAG", "Notes",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    for r_idx, r in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=r.period)
        ws.cell(row=r_idx, column=2, value=r.cost_centre)
        ws.cell(row=r_idx, column=3, value=r.category)
        ws.cell(row=r_idx, column=4, value=r.line_type)
        bud = ws.cell(row=r_idx, column=5, value=r.budget)
        bud.number_format = GBP_FORMAT
        act = ws.cell(row=r_idx, column=6, value=r.actual)
        act.number_format = GBP_FORMAT
        var = ws.cell(row=r_idx, column=7, value=r.variance)
        var.number_format = GBP_FORMAT
        if r.variance_pct is not None:
            cell = ws.cell(row=r_idx, column=8, value=r.variance_pct)
            cell.number_format = PCT_FORMAT
        else:
            ws.cell(row=r_idx, column=8, value="n/a")
        tol = ws.cell(row=r_idx, column=9, value=r.tolerance_pct)
        tol.number_format = PCT_FORMAT
        ws.cell(row=r_idx, column=10, value="Yes" if r.is_adverse else "No")
        rag_cell = ws.cell(row=r_idx, column=11, value=r.rag.upper())
        rag_cell.alignment = CENTER
        rag_cell.font = Font(bold=True, color="FFFFFF")
        fill = RAG_FILL.get(r.rag)
        if fill is not None:
            rag_cell.fill = fill
        ws.cell(row=r_idx, column=12, value=r.notes or "")

    if rows:
        ws.auto_filter.ref = ws.dimensions

    widths = [12, 22, 22, 10, 14, 14, 14, 12, 12, 9, 12, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Commentary sheet -----------------------------------------------

def _write_commentary(ws, c, metadata) -> None:
    ws["A1"] = "AI Management Commentary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    sub = ws.cell(row=2, column=1,
                  value=f"{metadata.company}  |  {metadata.period_label}  |  {c.model or '(no model)'}")
    sub.font = Font(italic=True, color="555555")
    ws.merge_cells("A2:B2")

    if c.error:
        ws.cell(row=4, column=1, value="Status").font = Font(bold=True)
        ws.cell(row=4, column=2, value=f"AI commentary unavailable. {c.error}")
        ws.cell(row=4, column=2).alignment = LEFT
        ws.row_dimensions[4].height = 50
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 100
        return

    if c.skipped:
        ws.cell(row=4, column=1, value="Status").font = Font(bold=True)
        ws.cell(row=4, column=2, value="AI commentary skipped on this run.")
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 100
        return

    row = 4
    ws.cell(row=row, column=1, value="Headline").font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value=c.headline)
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 30
    row += 2

    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value=c.summary)
    cell.alignment = LEFT
    ws.row_dimensions[row].height = max(60, 18 * (1 + c.summary.count(". ")))
    row += 2

    ws.cell(row=row, column=1, value="Adverse drivers").font = Font(bold=True, color="C62828")
    text = "\n".join(f"- {d}" for d in c.adverse_drivers) or "(none)"
    cell = ws.cell(row=row, column=2, value=text)
    cell.alignment = LEFT
    ws.row_dimensions[row].height = max(40, 18 * max(len(c.adverse_drivers), 1))
    row += 2

    ws.cell(row=row, column=1, value="Favourable drivers").font = Font(bold=True, color="1F4E79")
    text = "\n".join(f"- {d}" for d in c.favourable_drivers) or "(none)"
    cell = ws.cell(row=row, column=2, value=text)
    cell.alignment = LEFT
    ws.row_dimensions[row].height = max(40, 18 * max(len(c.favourable_drivers), 1))
    row += 2

    ws.cell(row=row, column=1, value="Actions").font = Font(bold=True, color="2E8540")
    text = "\n".join(f"- {a}" for a in c.actions) or "(none)"
    cell = ws.cell(row=row, column=2, value=text)
    cell.alignment = LEFT
    ws.row_dimensions[row].height = max(40, 18 * max(len(c.actions), 1))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


# ---- Inputs sheet ---------------------------------------------------

def _write_inputs(ws, result) -> None:
    headers = [
        "Period", "Cost Centre", "Category", "Line type",
        "Budget", "Actual", "Tolerance %", "Notes",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    for r_idx, r in enumerate(result.rows, start=2):
        ws.cell(row=r_idx, column=1, value=r.period)
        ws.cell(row=r_idx, column=2, value=r.cost_centre)
        ws.cell(row=r_idx, column=3, value=r.category)
        ws.cell(row=r_idx, column=4, value=r.line_type)
        bud = ws.cell(row=r_idx, column=5, value=r.budget)
        bud.number_format = GBP_FORMAT
        act = ws.cell(row=r_idx, column=6, value=r.actual)
        act.number_format = GBP_FORMAT
        tol = ws.cell(row=r_idx, column=7, value=r.tolerance_pct)
        tol.number_format = PCT_FORMAT
        ws.cell(row=r_idx, column=8, value=r.notes or "")

    widths = [12, 22, 22, 10, 14, 14, 12, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Helpers --------------------------------------------------------

def _slug(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "untitled"
