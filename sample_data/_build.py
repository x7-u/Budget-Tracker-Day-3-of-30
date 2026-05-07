"""Day 3. One-shot generator for the bundled budget vs actual samples.

Three samples are produced:

* sample_general.xlsx     6 cost centres x 5 categories, single month, all cost lines
* sample_marketing.xlsx   4 sub-teams x 6 categories, single month, per-row tolerance overrides
* sample_quarterly.xlsx   3 quarters x 4 cost centres x 4 categories, mixed cost and revenue lines

Each sample is deterministic (seeded random) and asserts that it round-trips
through ``parse_inputs`` and produces the expected variance row count before
saving. Run ``python _build.py`` to regenerate the bundled XLSX files.
"""
from __future__ import annotations

import random
import sys
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).resolve().parent
DAY_ROOT = HERE.parent
PROJECT_ROOT = DAY_ROOT.parent
for p in (str(DAY_ROOT), str(PROJECT_ROOT)):
    if p not in sys.path:
        sys.path.insert(0, p)

from budget_schema import parse_inputs  # noqa: E402

DATA_HEADERS = (
    "period", "cost_centre", "category", "line_type",
    "budget", "actual", "tolerance_pct", "notes",
    "forecast", "parent",
)


# ---- Generic helpers --------------------------------------------------

def _new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _write_metadata(
    wb: Workbook,
    *,
    company: str,
    fiscal_year: str,
    period_label: str,
) -> None:
    ws = wb.create_sheet("metadata")
    ws.append(["key", "value"])
    ws.append(["company", company])
    ws.append(["currency", "GBP"])
    ws.append(["fiscal_year", fiscal_year])
    ws.append(["period_label", period_label])


def _write_data_long(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("data")
    ws.append(list(DATA_HEADERS))
    for r in rows:
        ws.append([r.get(c) for c in DATA_HEADERS])


def _save_and_verify(wb: Workbook, path: Path, *, expected_rows: int) -> None:
    wb.save(path)
    parsed = parse_inputs(path=path)
    assert len(parsed.rows) == expected_rows, (
        f"{path.name}: expected {expected_rows} rows after parsing, "
        f"got {len(parsed.rows)}"
    )
    print(f"  wrote {path.name}  ({len(parsed.rows)} rows)")


def _round_amount(v: float) -> float:
    return round(v, 2)


# ---- Sample 1: General SME -------------------------------------------

def build_general() -> Path:
    rng = random.Random(42)
    cost_centres = ["Engineering", "Product", "Operations", "Sales", "Marketing", "Finance"]
    categories = ["Salaries", "Software", "Travel", "Office", "Professional Fees"]

    rows: list[dict] = []
    for cc in cost_centres:
        for cat in categories:
            base = _budget_for(cc, cat)
            # Generate a controlled mix of green / amber / red bands.
            band = rng.choice(["green", "green", "amber", "amber", "red", "favourable"])
            actual = _amount_for_band(base, band, rng)
            rows.append({
                "period": "2026-04",
                "cost_centre": cc,
                "category": cat,
                "line_type": "cost",
                "budget": base,
                "actual": actual,
                "tolerance_pct": None,
                "notes": None,
            })

    wb = _new_workbook()
    _write_metadata(wb, company="Sample Co Ltd", fiscal_year="FY26", period_label="April 2026")
    _write_data_long(wb, rows)
    out = HERE / "sample_general.xlsx"
    _save_and_verify(wb, out, expected_rows=len(rows))
    return out


def _budget_for(cost_centre: str, category: str) -> float:
    base_by_cc = {
        "Engineering": 120_000,
        "Product":     95_000,
        "Operations":  78_000,
        "Sales":       110_000,
        "Marketing":   85_000,
        "Finance":     65_000,
    }
    cat_factor = {
        "Salaries":          0.70,
        "Software":          0.10,
        "Travel":            0.05,
        "Office":            0.07,
        "Professional Fees": 0.08,
    }
    return _round_amount(base_by_cc[cost_centre] * cat_factor[category])


def _amount_for_band(budget: float, band: str, rng: random.Random) -> float:
    """Produce an actual that lands in the requested RAG band on a cost line."""
    if band == "green":
        delta = rng.uniform(-0.04, 0.045)
    elif band == "amber":
        delta = rng.uniform(0.06, 0.14)
    elif band == "red":
        delta = rng.uniform(0.16, 0.30)
    else:  # favourable, deeper underspend
        delta = rng.uniform(-0.20, -0.07)
    return _round_amount(budget * (1 + delta))


# ---- Sample 2: Marketing function deep-dive -------------------------

def build_marketing() -> Path:
    rng = random.Random(7)
    teams = ["Brand", "Digital", "Events", "PR"]
    categories = ["Salaries", "Agency Fees", "Paid Media", "Travel", "Tools", "Events"]

    rows: list[dict] = []
    for team in teams:
        for cat in categories:
            base = _marketing_budget(team, cat)
            band = rng.choice(["green", "amber", "amber", "red", "favourable"])
            actual = _amount_for_band(base, band, rng)
            tolerance = 0.15 if (team == "Events" and cat == "Travel") else None
            rows.append({
                "period": "2026-04",
                "cost_centre": team,
                "category": cat,
                "line_type": "cost",
                "budget": base,
                "actual": actual,
                "tolerance_pct": tolerance,
                "notes": ("Conference seasonality, wider tolerance band"
                          if tolerance is not None else None),
            })

    wb = _new_workbook()
    _write_metadata(wb, company="Acme Marketing", fiscal_year="FY26", period_label="April 2026")
    _write_data_long(wb, rows)
    out = HERE / "sample_marketing.xlsx"
    _save_and_verify(wb, out, expected_rows=len(rows))
    return out


def _marketing_budget(team: str, category: str) -> float:
    team_base = {"Brand": 45_000, "Digital": 78_000, "Events": 62_000, "PR": 32_000}
    cat_factor = {
        "Salaries":     0.50,
        "Agency Fees":  0.15,
        "Paid Media":   0.20,
        "Travel":       0.05,
        "Tools":        0.05,
        "Events":       0.05,
    }
    return _round_amount(team_base[team] * cat_factor[category])


# ---- Sample 3: Quarterly with revenue lines -------------------------

PARENT_BY_CC = {
    "Platform": "Engineering",
    "GTM": "Commercial",
    "Customer Success": "Commercial",
    "Corporate": "Corporate",
}


def build_quarterly() -> Path:
    rng = random.Random(2026)
    periods = ["2026-Q1", "2026-Q2", "2026-Q3"]
    cost_centres = ["Platform", "GTM", "Customer Success", "Corporate"]
    cost_categories = ["Salaries", "Cloud Infrastructure", "Travel", "Software"]

    rows: list[dict] = []
    # Cost lines, steady against plan. Forecast is set to budget by default
    # except for Cloud Infrastructure where Finance has revised expectations.
    for period in periods:
        for cc in cost_centres:
            for cat in cost_categories:
                base = _quarterly_cost_budget(cc, cat)
                band = rng.choice(["green", "green", "amber", "favourable"])
                actual = _amount_for_band(base, band, rng)
                forecast = base * 1.08 if cat == "Cloud Infrastructure" else base
                rows.append({
                    "period": period,
                    "cost_centre": cc,
                    "category": cat,
                    "line_type": "cost",
                    "budget": base,
                    "actual": actual,
                    "tolerance_pct": None,
                    "notes": None,
                    "forecast": _round_amount(forecast),
                    "parent": PARENT_BY_CC.get(cc),
                })

    # Revenue lines, ahead in Q1 then sliding.
    revenue_streams = [
        ("Platform", "Subscription Revenue"),
        ("GTM",      "Professional Services"),
    ]
    for period in periods:
        for cc, cat in revenue_streams:
            base = _quarterly_revenue_budget(cc, cat)
            actual = _revenue_actual_for_period(base, period, cc, rng)
            rows.append({
                "period": period,
                "cost_centre": cc,
                "category": cat,
                "line_type": "revenue",
                "budget": base,
                "actual": actual,
                "tolerance_pct": None,
                "notes": None,
                "forecast": _round_amount(base * 0.95),
                "parent": PARENT_BY_CC.get(cc),
            })

    wb = _new_workbook()
    _write_metadata(wb, company="TechStartup Ltd", fiscal_year="FY26",
                    period_label="Q1 to Q3 2026")
    _write_data_long(wb, rows)
    out = HERE / "sample_quarterly.xlsx"
    _save_and_verify(wb, out, expected_rows=len(rows))
    return out


def _quarterly_cost_budget(cc: str, cat: str) -> float:
    cc_base = {"Platform": 220_000, "GTM": 165_000, "Customer Success": 90_000, "Corporate": 70_000}
    cat_factor = {
        "Salaries":             0.60,
        "Cloud Infrastructure": 0.20,
        "Travel":               0.07,
        "Software":             0.13,
    }
    return _round_amount(cc_base[cc] * cat_factor[cat])


def _quarterly_revenue_budget(cc: str, cat: str) -> float:
    return {
        ("Platform", "Subscription Revenue"):    480_000.0,
        ("GTM",      "Professional Services"):    95_000.0,
    }[(cc, cat)]


def _revenue_actual_for_period(budget: float, period: str, cc: str, rng: random.Random) -> float:
    """Story arc: revenue ahead of plan in Q1, slipping by Q3."""
    delta_by_period = {
        "2026-Q1":  rng.uniform(0.07, 0.14),     # favourable
        "2026-Q2":  rng.uniform(-0.03, 0.04),    # green
        "2026-Q3":  rng.uniform(-0.18, -0.10),   # red
    }
    return _round_amount(budget * (1 + delta_by_period[period]))


# ---- Entry point -----------------------------------------------------

def build_all() -> None:
    print("Building Day 3 sample data...")
    build_general()
    build_marketing()
    build_quarterly()
    print("Done.")


if __name__ == "__main__":
    build_all()
