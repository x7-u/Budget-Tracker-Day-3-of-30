"""Day 3. Schema, validators and parser for the budget vs actual workbook.

Three input shapes are accepted:

1. Long format XLSX. One sheet called ``data`` with the columns
   ``period | cost_centre | category | line_type | budget | actual``
   plus optional ``tolerance_pct``, ``notes``, ``forecast`` and ``parent``.

2. Wide format XLSX. Two sheets called ``budget`` and ``actual`` with the
   columns ``period | cost_centre | category | line_type | amount``.
   The two sheets are inner-joined on the four-key tuple to produce the
   canonical long shape.

3. CSV. A flat file with the same columns as the long-format ``data``
   sheet. Metadata is derived from sensible defaults (filename for the
   company, GBP for currency, current year for fiscal year). Use this
   path when you want to paste a query result without building a full
   workbook.

A workbook with both shapes 1 and 2 present is rejected. ``line_type``
defaults to ``cost`` when missing. The MVP only supports GBP.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_DATA_COLS = ("period", "cost_centre", "category", "budget", "actual")
OPTIONAL_DATA_COLS = ("line_type", "tolerance_pct", "notes", "forecast", "parent")
DATA_COLS = REQUIRED_DATA_COLS + OPTIONAL_DATA_COLS

REQUIRED_WIDE_COLS = ("period", "cost_centre", "category", "amount")
OPTIONAL_WIDE_COLS = ("line_type", "parent")
WIDE_COLS = REQUIRED_WIDE_COLS + OPTIONAL_WIDE_COLS

REQUIRED_METADATA_KEYS = ("company", "currency", "fiscal_year", "period_label")
ALLOWED_LINE_TYPES = ("cost", "revenue")
SUPPORTED_CURRENCIES = ("GBP",)
MAX_ROWS = 1000


@dataclass(frozen=True)
class BudgetMetadata:
    company: str
    currency: str
    fiscal_year: str
    period_label: str


@dataclass
class BudgetRow:
    period: str
    cost_centre: str
    category: str
    line_type: str
    budget: float
    actual: float
    tolerance_pct: float | None = None
    notes: str | None = None
    forecast: float | None = None    # optional third value (current period forecast)
    parent: str | None = None        # optional parent cost centre for hierarchy rollups


@dataclass
class ParsedInputs:
    metadata: BudgetMetadata
    rows: list[BudgetRow]
    warnings: list[str] = field(default_factory=list)


# ---- Public entry point ------------------------------------------------

def parse_inputs(
    *,
    file_bytes: bytes | None = None,
    path: Path | str | None = None,
    source_filename: str | None = None,
) -> ParsedInputs:
    """Read an XLSX or CSV input and produce a normalised ParsedInputs.

    Either ``file_bytes`` or ``path`` must be provided. ``source_filename``
    is used to detect the file type and to derive the company name when
    parsing a CSV that has no metadata sheet.
    """
    if file_bytes is None and path is None:
        raise ValueError("parse_inputs needs either file_bytes or path")

    detect_name = source_filename or (str(path) if path is not None else "")
    suffix = Path(detect_name).suffix.lower()

    if suffix == ".csv":
        return _parse_csv_inputs(file_bytes=file_bytes, path=path,
                                 source_filename=source_filename)

    if file_bytes is not None:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)
    else:
        wb = load_workbook(filename=str(path), data_only=True, read_only=False)

    sheet_names = {n.lower(): n for n in wb.sheetnames}
    metadata_name = sheet_names.get("metadata")
    if metadata_name is None:
        raise ValueError(
            "Workbook is missing a 'metadata' sheet. "
            "Expected keys: company, currency, fiscal_year, period_label."
        )
    metadata = _parse_metadata(wb[metadata_name])

    has_data  = "data" in sheet_names
    has_budget = "budget" in sheet_names
    has_actual = "actual" in sheet_names

    if has_data and (has_budget or has_actual):
        raise ValueError(
            "Workbook contains both layouts. Provide either a 'data' sheet "
            "or 'budget' plus 'actual' sheets, not both."
        )

    warnings: list[str] = []

    if has_data:
        rows = _parse_long(wb[sheet_names["data"]])
    elif has_budget and has_actual:
        rows = _parse_wide(wb[sheet_names["budget"]], wb[sheet_names["actual"]])
    else:
        if has_budget and not has_actual:
            raise ValueError(
                "Workbook has a 'budget' sheet but no 'actual' sheet. "
                "Wide format requires both. Add an 'actual' sheet or switch to a single 'data' sheet."
            )
        if has_actual and not has_budget:
            raise ValueError(
                "Workbook has an 'actual' sheet but no 'budget' sheet. "
                "Wide format requires both. Add a 'budget' sheet or switch to a single 'data' sheet."
            )
        raise ValueError(
            "Workbook is missing the data sheets. Provide either a 'data' sheet "
            "(long format) or both 'budget' and 'actual' sheets (wide format)."
        )

    rows, dropped = _drop_all_zero(rows)
    if dropped:
        warnings.append(f"Dropped {dropped} empty row(s) (budget and actual both zero).")

    rows, capped = _cap_rows(rows)
    if capped:
        warnings.append(f"Row count exceeded {MAX_ROWS}; truncated to the first {MAX_ROWS} rows.")

    _check_unique(rows)
    warnings.extend(_check_in_progress(rows))
    warnings.extend(_check_revenue_signs(rows))
    return ParsedInputs(metadata=metadata, rows=rows, warnings=warnings)


# ---- Metadata ----------------------------------------------------------

def _parse_metadata(ws) -> BudgetMetadata:
    """Metadata sheet is two columns: key in column A, value in column B.

    Header row optional. Keys are case-insensitive, trimmed.
    """
    pairs: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        key = _as_str(row[0]).strip().lower()
        if not key or key in {"key", "field", "name"}:
            continue
        val = _as_str(row[1] if len(row) > 1 else "").strip()
        pairs[key] = val

    missing = [k for k in REQUIRED_METADATA_KEYS if not pairs.get(k)]
    if missing:
        raise ValueError(
            f"metadata sheet is missing required key(s): {', '.join(missing)}. "
            f"Expected: {', '.join(REQUIRED_METADATA_KEYS)}."
        )

    currency = pairs["currency"].upper()
    if currency not in SUPPORTED_CURRENCIES:
        raise ValueError(
            f"Day 3 MVP supports GBP only. Found currency '{currency}'. "
            "Multi-currency support is on the post-30 backlog."
        )

    return BudgetMetadata(
        company=pairs["company"],
        currency=currency,
        fiscal_year=pairs["fiscal_year"],
        period_label=pairs["period_label"],
    )


# ---- Long format -------------------------------------------------------

def _parse_long(ws) -> list[BudgetRow]:
    headers, body = _read_table(ws)
    _check_required(headers, REQUIRED_DATA_COLS, sheet_name="data")
    idx = {col: headers.index(col) for col in headers if col in DATA_COLS}

    rows: list[BudgetRow] = []
    for line_no, row in enumerate(body, start=2):
        if all(_is_blank(v) for v in row):
            continue
        try:
            br = _build_row(row, idx, line_no, sheet_name="data")
        except ValueError:
            raise
        rows.append(br)
    return rows


# ---- Wide format -------------------------------------------------------

def _parse_wide(ws_budget, ws_actual) -> list[BudgetRow]:
    bud_rows = _read_wide_sheet(ws_budget, sheet_name="budget")
    act_rows = _read_wide_sheet(ws_actual, sheet_name="actual")

    actual_lookup: dict[tuple[str, str, str, str], float] = {}
    for r in act_rows:
        key = (r["period"], r["cost_centre"], r["category"], r["line_type"])
        if key in actual_lookup:
            raise ValueError(f"Duplicate row in 'actual' sheet for key {key}.")
        actual_lookup[key] = r["amount"]

    seen: set[tuple[str, str, str, str]] = set()
    rows: list[BudgetRow] = []
    for r in bud_rows:
        key = (r["period"], r["cost_centre"], r["category"], r["line_type"])
        if key in seen:
            raise ValueError(f"Duplicate row in 'budget' sheet for key {key}.")
        seen.add(key)
        actual = actual_lookup.get(key, 0.0)
        rows.append(BudgetRow(
            period=r["period"],
            cost_centre=r["cost_centre"],
            category=r["category"],
            line_type=r["line_type"],
            budget=r["amount"],
            actual=actual,
            tolerance_pct=None,
            notes=None,
            forecast=None,
            parent=r.get("parent"),
        ))
    return rows


def _read_wide_sheet(ws, *, sheet_name: str) -> list[dict]:
    headers, body = _read_table(ws)
    _check_required(headers, REQUIRED_WIDE_COLS, sheet_name=sheet_name)
    idx = {col: headers.index(col) for col in headers if col in WIDE_COLS}
    out: list[dict] = []
    for line_no, row in enumerate(body, start=2):
        if all(_is_blank(v) for v in row):
            continue
        out.append({
            "period":      _as_str(row[idx["period"]]).strip(),
            "cost_centre": _as_str(row[idx["cost_centre"]]).strip(),
            "category":    _as_str(row[idx["category"]]).strip(),
            "line_type":   _read_line_type(row, idx, line_no=line_no, sheet_name=sheet_name),
            "amount":      _as_number(row[idx["amount"]], col="amount", line_no=line_no, sheet_name=sheet_name),
            "parent":      _read_optional_str(row, idx, "parent"),
        })
    return out


# ---- Row builders ------------------------------------------------------

def _build_row(row, idx, line_no: int, *, sheet_name: str) -> BudgetRow:
    period = _as_str(row[idx["period"]]).strip()
    cost_centre = _as_str(row[idx["cost_centre"]]).strip()
    category = _as_str(row[idx["category"]]).strip()
    if not (period and cost_centre and category):
        raise ValueError(
            f"{sheet_name}!row {line_no}: period, cost_centre and category are all required."
        )

    budget_raw = row[idx["budget"]]
    actual_raw = row[idx["actual"]]
    if _is_blank(budget_raw):
        raise ValueError(
            f"{sheet_name}!row {line_no}: 'budget' is missing for "
            f"{period} / {cost_centre} / {category}."
        )

    return BudgetRow(
        period=period,
        cost_centre=cost_centre,
        category=category,
        line_type=_read_line_type(row, idx, line_no=line_no, sheet_name=sheet_name),
        budget=_as_number(budget_raw, col="budget", line_no=line_no, sheet_name=sheet_name),
        actual=(0.0 if _is_blank(actual_raw)
                else _as_number(actual_raw, col="actual", line_no=line_no, sheet_name=sheet_name)),
        tolerance_pct=_read_optional_pct(row, idx, "tolerance_pct"),
        notes=_read_optional_str(row, idx, "notes"),
        forecast=_read_optional_number(row, idx, "forecast"),
        parent=_read_optional_str(row, idx, "parent"),
    )


def _read_line_type(row, idx, *, line_no: int, sheet_name: str) -> str:
    if "line_type" not in idx:
        return "cost"
    raw = _as_str(row[idx["line_type"]]).strip().lower()
    if not raw:
        return "cost"
    if raw not in ALLOWED_LINE_TYPES:
        raise ValueError(
            f"{sheet_name}!row {line_no}: line_type must be one of "
            f"{ALLOWED_LINE_TYPES}, got '{raw}'."
        )
    return raw


def _read_optional_pct(row, idx, name: str) -> float | None:
    if name not in idx:
        return None
    v = row[idx[name]]
    if _is_blank(v):
        return None
    return float(v)


def _read_optional_number(row, idx, name: str) -> float | None:
    if name not in idx:
        return None
    v = row[idx[name]]
    if _is_blank(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("£", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _read_optional_str(row, idx, name: str) -> str | None:
    if name not in idx:
        return None
    v = _as_str(row[idx[name]]).strip()
    return v or None


# ---- Sheet reading + cleaners -----------------------------------------

def _read_table(ws) -> tuple[list[str], list[tuple]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{ws.title}' is empty.")
    headers_raw = rows[0]
    headers = [_as_str(c).strip().lower() for c in headers_raw]
    body = rows[1:]
    return headers, body


def _check_required(headers: list[str], required: tuple[str, ...], *, sheet_name: str) -> None:
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' is missing column(s): {', '.join(missing)}. "
            f"Expected at minimum: {', '.join(required)}."
        )


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    return str(v)


def _as_number(v: Any, *, col: str, line_no: int, sheet_name: str) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("£", "")
        try:
            return float(s)
        except ValueError as e:
            raise ValueError(
                f"{sheet_name}!row {line_no}: '{col}' must be numeric, got '{v}'."
            ) from e
    raise ValueError(
        f"{sheet_name}!row {line_no}: '{col}' must be numeric, got {type(v).__name__}."
    )


# ---- Cross-row validators ---------------------------------------------

def _drop_all_zero(rows: list[BudgetRow]) -> tuple[list[BudgetRow], int]:
    kept: list[BudgetRow] = []
    dropped = 0
    for r in rows:
        if r.budget == 0 and r.actual == 0:
            dropped += 1
            continue
        kept.append(r)
    return kept, dropped


def _cap_rows(rows: list[BudgetRow]) -> tuple[list[BudgetRow], bool]:
    if len(rows) <= MAX_ROWS:
        return rows, False
    return rows[:MAX_ROWS], True


def _check_unique(rows: list[BudgetRow]) -> None:
    seen: dict[tuple[str, str, str, str], int] = {}
    for i, r in enumerate(rows, start=1):
        key = (r.period, r.cost_centre, r.category, r.line_type)
        if key in seen:
            raise ValueError(
                f"Duplicate row: {key} appears at row {seen[key]} and row {i}."
            )
        seen[key] = i


def _check_in_progress(rows: list[BudgetRow]) -> list[str]:
    out: list[str] = []
    for r in rows:
        if r.budget != 0 and r.actual == 0:
            out.append(
                f"In-progress: {r.period}/{r.cost_centre}/{r.category} has no actual yet."
            )
    return out


def _check_revenue_signs(rows: list[BudgetRow]) -> list[str]:
    out: list[str] = []
    for r in rows:
        if r.line_type == "revenue" and (r.budget < 0 or r.actual < 0):
            out.append(
                f"Revenue line with a negative value at "
                f"{r.period}/{r.cost_centre}/{r.category}. "
                "Verify sign convention; not auto-flipped."
            )
    return out


# ---- CSV input -------------------------------------------------------

def _parse_csv_inputs(
    *,
    file_bytes: bytes | None,
    path: Path | str | None,
    source_filename: str | None,
) -> ParsedInputs:
    """Parse a flat CSV with the same columns as the long-format ``data``
    sheet. Metadata is derived: company from the filename (or "Imported
    CSV"), currency GBP, fiscal year from the current year."""
    if file_bytes is not None:
        text = file_bytes.decode("utf-8-sig", errors="replace")
    else:
        text = Path(path).read_text(encoding="utf-8-sig", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    headers_raw = reader.fieldnames or []
    headers = [h.strip().lower() for h in headers_raw]
    if not headers:
        raise ValueError("CSV is empty or has no header row.")
    missing = [c for c in REQUIRED_DATA_COLS if c not in headers]
    if missing:
        raise ValueError(
            f"CSV is missing column(s): {', '.join(missing)}. "
            f"Expected at minimum: {', '.join(REQUIRED_DATA_COLS)}."
        )

    # Map original header casing back so we can read with reader's keys.
    header_map = {h.strip().lower(): h for h in headers_raw}

    rows: list[BudgetRow] = []
    for line_no, raw in enumerate(reader, start=2):
        if all(_is_blank(raw.get(c)) for c in headers_raw):
            continue
        period = (raw.get(header_map["period"]) or "").strip()
        cost_centre = (raw.get(header_map["cost_centre"]) or "").strip()
        category = (raw.get(header_map["category"]) or "").strip()
        if not (period and cost_centre and category):
            raise ValueError(
                f"CSV row {line_no}: period, cost_centre and category are all required."
            )
        budget_raw = raw.get(header_map["budget"])
        actual_raw = raw.get(header_map["actual"])
        if _is_blank(budget_raw):
            raise ValueError(
                f"CSV row {line_no}: 'budget' is missing for "
                f"{period} / {cost_centre} / {category}."
            )

        line_type_raw = (raw.get(header_map["line_type"]) or "").strip().lower() if "line_type" in header_map else ""
        if not line_type_raw:
            line_type = "cost"
        elif line_type_raw not in ALLOWED_LINE_TYPES:
            raise ValueError(
                f"CSV row {line_no}: line_type must be one of "
                f"{ALLOWED_LINE_TYPES}, got '{line_type_raw}'."
            )
        else:
            line_type = line_type_raw

        rows.append(BudgetRow(
            period=period,
            cost_centre=cost_centre,
            category=category,
            line_type=line_type,
            budget=_as_number(budget_raw, col="budget", line_no=line_no, sheet_name="data"),
            actual=(0.0 if _is_blank(actual_raw)
                    else _as_number(actual_raw, col="actual", line_no=line_no, sheet_name="data")),
            tolerance_pct=_csv_pct(raw, header_map, "tolerance_pct"),
            notes=_csv_str(raw, header_map, "notes"),
            forecast=_csv_number(raw, header_map, "forecast"),
            parent=_csv_str(raw, header_map, "parent"),
        ))

    name_hint = source_filename or (str(path) if path is not None else "")
    metadata = BudgetMetadata(
        company=_company_from_filename(name_hint),
        currency="GBP",
        fiscal_year=f"FY{dt.date.today().year % 100:02d}",
        period_label="Imported CSV",
    )

    warnings: list[str] = ["Metadata defaults applied because the input was a CSV. "
                          "For full control, supply an XLSX with a metadata sheet."]
    rows, dropped = _drop_all_zero(rows)
    if dropped:
        warnings.append(f"Dropped {dropped} empty row(s) (budget and actual both zero).")
    rows, capped = _cap_rows(rows)
    if capped:
        warnings.append(f"Row count exceeded {MAX_ROWS}; truncated to the first {MAX_ROWS} rows.")

    _check_unique(rows)
    warnings.extend(_check_in_progress(rows))
    warnings.extend(_check_revenue_signs(rows))
    return ParsedInputs(metadata=metadata, rows=rows, warnings=warnings)


def _csv_pct(raw, header_map, name):
    if name not in header_map:
        return None
    v = raw.get(header_map[name])
    if _is_blank(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _csv_number(raw, header_map, name):
    if name not in header_map:
        return None
    v = raw.get(header_map[name])
    if _is_blank(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("£", "")
    try:
        return float(s)
    except ValueError:
        return None


def _csv_str(raw, header_map, name):
    if name not in header_map:
        return None
    v = (raw.get(header_map[name]) or "").strip()
    return v or None


def _company_from_filename(name: str) -> str:
    if not name:
        return "Imported CSV"
    stem = Path(name).stem.replace("_", " ").replace("-", " ").strip()
    return stem.title() if stem else "Imported CSV"
